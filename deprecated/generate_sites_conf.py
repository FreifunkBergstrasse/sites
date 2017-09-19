#!/usr/local/bin/python3

# todo: usage, parameters (excelfile, output-directory), error handling
# todo: version for python 2.7?
# todo: check template for correct syntax and containing placeholders
# todo: check excel to contain a headline and at least one line of data; each row must have same number of columns

# parameter: <excel> <source-dir> <target-dir> -verbose
# allow usage of row contents to be used as part of directory and/or filename
# allow dry runs

from openpyxl import load_workbook
import re
# import sys
# print (sys.argv)

# create empty dictionary
dict = {}

# define regular expression to match placeholders (= '@' + any character or number or underscore or dot + '@')
placeHolderRegEx = re.compile(r'@[a-zA-Z0-9_.]+@')

# open excel sheet with individual data for each domain
wb = load_workbook('domains.xlsx', data_only=True)
sheet = wb.active


# read template site.conf into a string
templateFile = open('site.conf-template', 'r')
template = templateFile.read()
templateFile.close()

# extract all placeholders from template into a list, make the list unique (handle it as a set)
placeholders = set(placeHolderRegEx.findall(template))
print ('Replacing ' + str(len(placeholders)) + ' placeholders')
print ('Creating ' + str(sheet.max_row - 1) + ' files')

# iterate each row of the sheet; upper index not in range
for row in range(2, sheet.max_row +1):

	new = template

	#for row in sheet[1]:
	#	print(row.value)
	# fill dictionary with key value pairs from excel (headline = keys, line 1 to n = values); .row starts at 0
	# for (headline, data) in zip(sheet.rows[0], sheet.rows[row-1]):
	for (headline, data) in zip(sheet[1], sheet[row]):
		dict.update({'@'+headline.value+'@' : str(data.value)})

	# replace placeholders in template with the values fetched from the dictionary
	for placeholder in placeholders:
		new = new.replace(placeholder, dict[placeholder])

	# save new site.conf
	dir = dict['@Ort@'] + '/site'
	newFile = open(dir+'/site.conf', 'w')
	newFile.write(new)
	newFile.close()
